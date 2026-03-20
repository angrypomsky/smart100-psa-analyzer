# -*- coding: utf-8 -*-
"""
=============================================================================
SMART100 DET → ET 변환기  [일괄 처리 버전]
=============================================================================
사용법 (명령행):
  python et_generator.py                     # 현재 폴더의 모든 CSV 처리
  python et_generator.py data/               # 특정 폴더의 모든 CSV 처리
  python et_generator.py a.csv b.csv         # 개별 파일 지정

[사고유형별 헤딩 자동 감지]
  SLOCA2              → RT, PSIS, SIT (3헤딩)
  그 외 전부           → RT, PRHRS, ADS, PSIS, SIT (5헤딩)

[컬럼명 자동 정규화]
  PRHRS_HX_count → PRHRS_count
  RT_status      → Reactor_Trip
  RCP_status     → RCP_Status
  PCT_K          → PCT_max
  State          → Outcome
=============================================================================
"""

import sys
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ── 사고유형별 기본 헤딩 ────────────────────────────────────────────────
DEFAULT_HEADINGS = [
    'Reactor_Trip', 'PRHRS_count', 'ADS_BLEED_count',
    'PSIS_FEED_status', 'SIT_Refill_time'
]
HEADINGS_BY_TYPE = {
    'SLOCA2': ['Reactor_Trip', 'PSIS_FEED_status', 'SIT_Refill_time'],
}

# ── 컬럼명 정규화 맵 ────────────────────────────────────────────────────
RENAME_MAP = {
    'PRHRS_HX_count': 'PRHRS_count',
    'RT_status':       'Reactor_Trip',
    'RCP_status':      'RCP_Status',
    'PCT_K':           'PCT_max',
    'State':           'Outcome',
}

COL_DISPLAY = {
    'Reactor_Trip':     'Reactor Trip',
    'RCP_Status':       'RCP Status',
    'PRHRS_count':      'PRHRS 작동 계통수',
    'ADS_BLEED_count':  'ADS Bleed 작동 수',
    'PSIS_FEED_status': 'PSIS Feed 상태',
    'SIT_Refill_time':  'SIT Refill 시간',
    'PCT_max':          'PCT (K)',
    'Outcome':          'Outcome',
    'Seq_Num':          'Sequence 번호',
}


class ET_Generator:

    def __init__(self, output_dir: str = '.'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        print("=" * 70)
        print("SMART100 Event Tree 생성기  [일괄 처리 버전]")
        print("=" * 70)

    # ── 공개 API ──────────────────────────────────────────────────────
    def run_all(self, csv_paths: list, custom_headings: dict = None):
        """
        CSV 파일 목록을 받아 일괄 ET 생성

        Parameters
        ----------
        csv_paths : list[str | Path]
            처리할 CSV 파일 경로 목록
        custom_headings : dict (선택)
            사고유형별 헤딩 오버라이드
            예) {'SGTR': ['Reactor_Trip', 'PRHRS_count', 'ADS_BLEED_count']}
        """
        if not csv_paths:
            print("처리할 CSV 파일이 없습니다.")
            return

        results = {}
        errors  = []

        for csv_path in csv_paths:
            csv_path = Path(csv_path)
            print(f"\n{'─'*70}")
            print(f"처리 중: {csv_path.name}")
            print('─' * 70)
            try:
                output = self._process_single(
                    csv_path, custom_headings or {}
                )
                results[csv_path.name] = output
            except Exception as e:
                print(f"  오류: {e}")
                errors.append((csv_path.name, str(e)))

        # ── 결과 요약 ──
        print(f"\n{'='*70}")
        print(f"완료: {len(results)}개 성공 / {len(errors)}개 실패")
        for fname, out in results.items():
            print(f"   {fname} → {out}")
        if errors:
            print("\n실패 목록:")
            for fname, err in errors:
                print(f"   {fname}: {err}")
        print('=' * 70)

    # ── 단일 파일 처리 ────────────────────────────────────────────────
    def _process_single(self, csv_path: Path, custom_headings: dict):
        # 1) 로드 + 정규화 (인코딩 자동 감지)
        raw = None
        for enc in ('utf-8-sig', 'cp949', 'euc-kr', 'utf-8', 'latin-1'):
            try:
                raw = pd.read_csv(csv_path, encoding=enc)
                break
            except (UnicodeDecodeError, Exception):
                continue
        if raw is None:
            raise ValueError("CSV 인코딩을 감지할 수 없습니다.")
        df = self._normalize(raw)

        # 2) 시나리오 이름 + 헤딩 결정
        scenario_name = (
            csv_path.stem
                    .replace('_results', '')
                    .replace('__6_', '')
                    .upper()
        )
        headings = (
            custom_headings.get(scenario_name)
            or HEADINGS_BY_TYPE.get(scenario_name)
            or DEFAULT_HEADINGS
        )

        # 헤딩 존재 확인 → 없는 건 제외 후 경고
        valid_headings = []
        for h in headings:
            if h in df.columns:
                valid_headings.append(h)
            else:
                print(f"  경고: 헤딩 '{h}' 없음 → 건너뜀")
        headings = valid_headings

        print(f"  시나리오: {scenario_name}")
        print(f"  행 수: {len(df)}  |  헤딩: {headings}")

        # 3) 트리 생성
        branches = self._build_branches(df, headings)
        print(f"  분기 수: {len(branches)}")

        # 4) 결과 요약 출력
        total = len(df)
        for b in sorted(branches, key=lambda x: -x['count'])[:5]:
            path_str = ' → '.join(str(b[h]) for h in headings)
            cd = b.get('CD_count', 0)
            print(f"    {path_str}: {b['count']}개 ({b['count']/total*100:.1f}%) "
                  f"[CD:{cd}]")
        if len(branches) > 5:
            print(f"    ... 외 {len(branches)-5}개 분기")

        # 5) Excel 생성 (로컬 저장)
        output_file = self.output_dir / f'{scenario_name}_ET_result.xlsx'
        self._export(df, headings, branches, scenario_name, output_file)
        return str(output_file)

    # ── 데이터 정규화 ─────────────────────────────────────────────────
    def _normalize(self, df):
        df = df.rename(columns=RENAME_MAP)
        for c in ['SIT_Refill_time', 'PSIS_FEED_status', 'Reactor_Trip',
                  'RCP_Status', 'Outcome']:
            if c in df.columns:
                df[c] = df[c].fillna('N/A')
        return df

    # ── 트리 분기 생성 ────────────────────────────────────────────────
    def _build_branches(self, df, headings):
        tree = {}
        for idx, row in df.iterrows():
            path = tuple(row[h] for h in headings)
            tree.setdefault(path, []).append(idx)

        total = len(df)
        branches = []
        for path, ids in tree.items():
            b = {'path': path, 'count': len(ids),
                 'probability': len(ids) / total, 'scenario_ids': ids}
            for i, h in enumerate(headings):
                b[h] = path[i]
            if 'Outcome' in df.columns:
                outcomes = df.loc[ids, 'Outcome']
                b['CD_count'] = int((outcomes == 'CD').sum())
                b['OK_count'] = int((outcomes == 'OK').sum())
            branches.append(b)

        branches.sort(key=lambda x: x['probability'], reverse=True)
        return branches

    # ── Excel 생성 ────────────────────────────────────────────────────
    def _export(self, df, headings, branches, scenario_name, output_file):
        wb = Workbook()
        ws1 = wb.active; ws1.title = "원본_데이터"
        self._sheet_data(ws1, df, headings)

        ws2 = wb.create_sheet("분기_집계")
        self._sheet_branches(ws2, headings, branches, len(df))

        ws3 = wb.create_sheet("ET_구조")
        self._sheet_tree(ws3, headings, branches, scenario_name)

        ws4 = wb.create_sheet("통계")
        self._sheet_stats(ws4, df, headings, branches, scenario_name)

        wb.save(output_file)
        print(f"  저장 완료: {output_file}")

    # ── 시트: 원본_데이터 ─────────────────────────────────────────────
    def _sheet_data(self, ws, df, headings):
        priority = headings + ['PCT_max', 'PCT_pass', 'Outcome']
        export_cols = []
        for c in priority:
            if c in df.columns and c not in export_cols:
                export_cols.append(c)
        for c in df.columns:
            if c not in export_cols and c != 'Scenario_ID':
                export_cols.append(c)

        db_exp = df[[c for c in export_cols if c in df.columns]]
        hfill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hfont  = Font(bold=True, color="FFFFFF")
        for r_idx, row in enumerate(dataframe_to_rows(db_exp, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                if r_idx == 1:
                    cell.fill = hfill; cell.font = hfont
                    cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 30)

    # ── 시트: 분기_집계 ───────────────────────────────────────────────
    def _sheet_branches(self, ws, headings, branches, total):
        cols = ([COL_DISPLAY.get(h, h) for h in headings]
                + ['CD 개수', 'OK 개수', '시나리오 수', '확률', '누적확률'])
        hfill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF")
        for c_idx, name in enumerate(cols, 1):
            cell = ws.cell(1, c_idx, name)
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal='center')

        cumul = 0
        for r, b in enumerate(branches, 2):
            cumul += b['probability']
            for c_idx, h in enumerate(headings, 1):
                ws.cell(r, c_idx, b[h])
            n = len(headings)
            ws.cell(r, n+1, b.get('CD_count', '-'))
            ws.cell(r, n+2, b.get('OK_count', '-'))
            ws.cell(r, n+3, b['count'])
            ws.cell(r, n+4, b['probability']).number_format = '0.0%'
            ws.cell(r, n+5, cumul).number_format            = '0.0%'

        for i in range(1, len(cols)+1):
            ltr = chr(64+i) if i <= 26 else 'A'+chr(64+i-26)
            ws.column_dimensions[ltr].width = 20

    # ── 시트: ET_구조 ─────────────────────────────────────────────────
    def _sheet_tree(self, ws, headings, branches, scenario_name):
        ws.cell(1, 1, "Event Tree 구조").font = Font(bold=True, size=14)
        ws.cell(3, 1, f"IE: {scenario_name}").font = Font(bold=True)
        self._draw_tree(ws, 5, 2, branches, headings, 0)

    def _draw_tree(self, ws, start_row, col, branches, headings, h_idx):
        if h_idx >= len(headings):
            return start_row
        h = headings[h_idx]
        display = COL_DISPLAY.get(h, h)

        groups = {}
        for b in branches:
            groups.setdefault(b[h], []).append(b)
        reverse_sort = (h == 'PRHRS_count')
        try:
            keys = sorted(groups.keys(), key=float, reverse=reverse_sort)
        except (TypeError, ValueError):
            keys = list(groups.keys())

        row = start_row
        ws.cell(row, col, display).font = Font(bold=True)
        row += 1

        for val in keys:
            sub = groups[val]
            cnt  = sum(b['count'] for b in sub)
            prob = sum(b['probability'] for b in sub)
            label = f"{int(val)}계통" if h == 'PRHRS_count' else str(val)

            ws.cell(row, col, label).alignment = Alignment(indent=2)
            ws.cell(row, col+1, f"{cnt}개")
            ws.cell(row, col+2, prob).number_format = '0.0%'
            row += 1

            if h_idx + 1 < len(headings):
                row = self._draw_tree(ws, row, col+3, sub, headings, h_idx+1)
        return row

    # ── 시트: 통계 ────────────────────────────────────────────────────
    def _sheet_stats(self, ws, df, headings, branches, scenario_name):
        r = 1
        ws.cell(r, 1, "분석 통계").font = Font(bold=True, size=14); r += 2
        for label, val in [("사고 유형:", scenario_name),
                            ("총 시나리오 수:", len(df)),
                            ("총 분기 수:", len(branches))]:
            ws.cell(r, 1, label).font = Font(bold=True)
            ws.cell(r, 2, val); r += 1
        r += 1

        if 'Outcome' in df.columns:
            ws.cell(r, 1, "Outcome 분포").font = Font(bold=True, size=12); r += 1
            for val, cnt in df['Outcome'].value_counts().items():
                ws.cell(r, 2, str(val))
                ws.cell(r, 3, cnt)
                ws.cell(r, 4, cnt/len(df)).number_format = '0.0%'; r += 1
            r += 1

        ws.cell(r, 1, "헤딩별 분포").font = Font(bold=True, size=12); r += 1
        for h in headings:
            ws.cell(r, 1, f"[{COL_DISPLAY.get(h, h)}]").font = Font(bold=True); r += 1
            for val, cnt in df[h].value_counts().sort_index(ascending=(h != 'PRHRS_count')).items():
                label = f"{int(val)}계통:" if h == 'PRHRS_count' else f"{val}:"
                ws.cell(r, 2, label)
                ws.cell(r, 3, cnt)
                ws.cell(r, 4, cnt/len(df)).number_format = '0.0%'; r += 1
            r += 1


# ── CLI 진입점 ────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='SMART100 DET → ET 변환기 (일괄 처리)'
    )
    parser.add_argument(
        'inputs', nargs='*',
        help='CSV 파일 또는 폴더 경로 (생략 시 현재 폴더의 모든 CSV)'
    )
    parser.add_argument(
        '-o', '--output', default='.',
        help='Excel 결과 저장 폴더 (기본값: 현재 폴더)'
    )
    args = parser.parse_args()

    # CSV 파일 수집
    csv_files = []
    if not args.inputs:
        csv_files = sorted(Path('.').glob('*.csv'))
    else:
        for inp in args.inputs:
            p = Path(inp)
            if p.is_dir():
                csv_files.extend(sorted(p.glob('*.csv')))
            elif p.suffix.lower() == '.csv' and p.exists():
                csv_files.append(p)
            else:
                print(f"경고: '{inp}' 를 찾을 수 없거나 CSV 파일이 아닙니다.")

    if not csv_files:
        print("처리할 CSV 파일이 없습니다.")
        sys.exit(1)

    et_gen = ET_Generator(output_dir=args.output)
    et_gen.run_all(csv_files)


if __name__ == '__main__':
    main()
