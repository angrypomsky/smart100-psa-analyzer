# -*- coding: utf-8 -*-
"""
=============================================================================
ET result → PSA DET Dashboard 데이터 로드  [v2 - 로컬 실행 버전]
=============================================================================
사용법:
  python load_to_dashboard.py -t template.xlsx result1.xlsx result2.xlsx
  python load_to_dashboard.py -t template.xlsx results/   # 폴더 내 전체
  python load_to_dashboard.py -t template.xlsx -o output/ results/

[템플릿 Data 시트 컬럼 순서 (고정)]
  A: rt_state          → Reactor_Trip
  B: prhrs_hx_count    → PRHRS_count
  C: ads_bleed_count   → ADS_BLEED_count
  D: psis_feed_count   → PSIS_FEED_status → 숫자 변환 (Success=1, Fail=0, N/A='-')
  E: sit_refill_time   → SIT_Refill_time
  F: pct_k             → PCT_max
  G: pct_pass          → PCT_pass
  H: state             → Outcome
  (J열): LAST_ROW
=============================================================================
"""

import sys
import argparse
import os
import re
import zipfile
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from lxml import etree

# ── 상수 ──────────────────────────────────────────────────────────────
BLUE  = '4472C4'
RED   = 'FF0000'
CD_THRESHOLD = 1477

PCT_HIST_BINS = [400, 700, 1000, 1300, 1477, 1700, 2000, 2300]
N_BINS = len(PCT_HIST_BINS) - 1  # 7구간


def _calc_cd_bin_idx():
    for i in range(len(PCT_HIST_BINS) - 1):
        if PCT_HIST_BINS[i] >= CD_THRESHOLD:
            return i
    return N_BINS - 1

CD_BIN_IDX = _calc_cd_bin_idx()


# ── 차트 색상 패치 ─────────────────────────────────────────────────────
def _patch_chart_colors(xlsx_path):
    xlsx_path = str(xlsx_path)
    tmp = xlsx_path + '_tmp'
    os.makedirs(tmp, exist_ok=True)
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        z.extractall(tmp)

    ns_c = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
    ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    for root_dir, dirs, fnames in os.walk(tmp):
        for fname in fnames:
            if fname.startswith('chart') and fname.endswith('.xml'):
                cf = os.path.join(root_dir, fname)
                tree = etree.parse(cf)
                root = tree.getroot()
                for ser in root.findall(f'.//{{{ns_c}}}ser'):
                    for dpt in ser.findall(f'{{{ns_c}}}dPt'):
                        ser.remove(dpt)
                    sp_pr = ser.find(f'{{{ns_c}}}spPr')
                    insert_idx = (list(ser).index(sp_pr)
                                  if sp_pr is not None else len(list(ser)))
                    for i in range(N_BINS):
                        color = RED if i >= CD_BIN_IDX else BLUE
                        dpt_xml = (
                            f'<c:dPt xmlns:c="{ns_c}" xmlns:a="{ns_a}">'
                            f'<c:idx val="{i}"/>'
                            f'<c:invertIfNegative val="0"/>'
                            f'<c:spPr>'
                            f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>'
                            f'<a:ln><a:solidFill><a:srgbClr val="{color}"/></a:solidFill></a:ln>'
                            f'</c:spPr>'
                            f'</c:dPt>'
                        )
                        ser.insert(insert_idx + i, etree.fromstring(dpt_xml))
                tree.write(cf, xml_declaration=True,
                           encoding='UTF-8', standalone=True)

    out = xlsx_path + '_patched.xlsx'
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for root_dir, dirs, fnames in os.walk(tmp):
            for fname in fnames:
                fp = os.path.join(root_dir, fname)
                zout.write(fp, os.path.relpath(fp, tmp))
    shutil.rmtree(tmp)
    os.replace(out, xlsx_path)


# ── 컬럼 변환 헬퍼 ────────────────────────────────────────────────────
def _psis_to_count(val):
    if val in (None, '-', 'N/A', float('nan')):
        return '-'
    s = str(val).strip().lower()
    if s == 'success':
        return 1
    if s == 'fail':
        return 0
    return '-'


def _to_val(v):
    if v is None:
        return '-'
    try:
        import math
        if math.isnan(float(v)):
            return '-'
    except (TypeError, ValueError):
        pass
    return v


def _build_data_row(row_dict):
    def get(row, *keys):
        for k in keys:
            if k in row and row[k] not in (None,):
                return row[k]
        return '-'

    rt       = get(row_dict, 'Reactor_Trip', 'rt_state', 'RT_status', 'rt_status')
    prhrs    = get(row_dict, 'PRHRS_count', 'prhrs_hx_count', 'PRHRS_HX_count')
    ads      = get(row_dict, 'ADS_BLEED_count', 'ads_bleed_count')
    psis_raw = get(row_dict, 'PSIS_FEED_status', 'psis_feed_count', 'psis_feed_status')
    sit      = get(row_dict, 'SIT_Refill_time', 'sit_refill_time')
    pct      = get(row_dict, 'PCT_max', 'PCT_K', 'pct_k')
    pct_p    = get(row_dict, 'PCT_pass', 'pct_pass')
    state    = get(row_dict, 'Outcome', 'State', 'state')

    psis = _psis_to_count(psis_raw)

    try:
        pct_pass = 'Pass' if float(pct) < CD_THRESHOLD else 'Fail'
    except (ValueError, TypeError):
        pct_pass = _to_val(pct_p)

    return [
        _to_val(rt), _to_val(prhrs), _to_val(ads), psis,
        _to_val(sit), _to_val(pct), pct_pass, _to_val(state),
    ]


# ── Calc 수식 생성 ─────────────────────────────────────────────────────
def _calc_filter_formula(r, n):
    return (
        f'=AND('
        f'OR(Dashboard!B5="All",Data!A{r}=Dashboard!B5),'
        f'OR(Data!B{r}="-",AND(Data!B{r}>=Dashboard!B6,Data!B{r}<=Dashboard!B7)),'
        f'OR(Data!C{r}="-",AND(Data!C{r}>=Dashboard!B8,Data!C{r}<=Dashboard!B9)),'
        f'OR(Data!D{r}="-",AND(Data!D{r}>=Dashboard!B10,Data!D{r}<=Dashboard!B11)),'
        f'OR(Dashboard!B12="All",Data!E{r}=Dashboard!B12)'
        f')'
    )


# ── 단일 파일 처리 ────────────────────────────────────────────────────
def process_one(et_path: Path, dash_path: Path, output_dir: Path) -> str:
    accident_name = (
        et_path.stem
               .replace('_ET_result', '').replace('_ET_RESULT', '')
               .replace('_results', '').replace('_RESULTS', '')
               .upper()
    )
    output_name = output_dir / f'PSA_DET_Dashboard_{accident_name}_filled.xlsx'

    # ── ET result 로드 ──────────────────────────────────────────────
    wb_et = openpyxl.load_workbook(et_path)
    if '원본_데이터' not in wb_et.sheetnames:
        raise ValueError("ET result에 '원본_데이터' 시트가 없습니다.")

    ws_src = wb_et['원본_데이터']
    et_headers = [c.value for c in ws_src[1]]
    et_rows_raw = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(et_headers, row))
        et_rows_raw.append(_build_data_row(row_dict))
    print(f"  원본_데이터 로드 완료 ({len(et_rows_raw)}행)")

    n = len(et_rows_raw)

    # ── Dashboard 템플릿 로드 ────────────────────────────────────────
    wb = openpyxl.load_workbook(dash_path)
    ws_data = wb['Data']
    ws_dash = wb['Dashboard']
    ws_calc = wb['Calc']

    # ── Data 시트 색상 매핑 추출 (템플릿 원본에서) ──────────────────
    # 값 → fill 매핑: 열별로 추출 (None/빈 셀 제외)
    col_fill_map: dict[int, dict[str, object]] = {}
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
        for cell in row:
            if cell.value is None:
                continue
            if cell.fill and cell.fill.fill_type not in (None, 'none'):
                col = cell.column
                key = str(cell.value).strip()
                if col not in col_fill_map:
                    col_fill_map[col] = {}
                if key not in col_fill_map[col]:
                    col_fill_map[col][key] = copy(cell.fill)

    # ── Data 시트 채우기 ─────────────────────────────────────────────
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
        for cell in row:
            cell.value = None
            cell.fill = openpyxl.styles.PatternFill(fill_type=None)

    for i, row_data in enumerate(et_rows_raw):
        r = i + 2
        for j, val in enumerate(row_data):
            cell = ws_data.cell(r, j + 1, val)
            col = j + 1
            key = str(val).strip()
            if col in col_fill_map and key in col_fill_map[col]:
                cell.fill = copy(col_fill_map[col][key])

    for cell in ws_data[1]:
        if cell.value == 'LAST_ROW':
            ws_data.cell(2, cell.column, n + 1)
            break

    print(f"  Data 시트 완료 ({n}행) | 샘플: {et_rows_raw[0]}")

    # ── Dashboard 시트 업데이트 ──────────────────────────────────────
    ws_dash['A2'] = f'SMART100 - {accident_name}'
    ws_dash['B6']  = 0
    ws_dash['B7']  = 4
    ws_dash['B8']  = 0
    ws_dash['B9']  = 2
    ws_dash['B10'] = 0
    ws_dash['B11'] = 1
    ws_dash['B14'] = CD_THRESHOLD

    # ── Calc 시트 수식 확장 ──────────────────────────────────────────
    for row in ws_calc.iter_rows(min_row=2, max_row=ws_calc.max_row):
        for cell in row:
            cell.value = None

    B_RNG = f'B2:B{n+1}'
    C_RNG = f'C2:C{n+1}'
    E_RNG = f'E2:E{n+1}'

    for i in range(n):
        r = i + 2
        ws_calc.cell(r, 1).value = r
        ws_calc.cell(r, 2).value = _calc_filter_formula(r, n)
        ws_calc.cell(r, 3).value = f'=IF(B{r},Data!F{r},"")'
        ws_calc.cell(r, 4).value = f'=IF(AND(B{r},Data!F{r}>=Dashboard!B14),1,0)'
        ws_calc.cell(r, 5).value = f'=IF(B{r},C{r}^2,"")'
        if i == 0:
            ws_calc.cell(r, 6).value  = f'=AVERAGEIF({B_RNG},TRUE(),{C_RNG})'
            ws_calc.cell(r, 7).value  = (
                f'=IFERROR(SQRT(AVERAGEIF({B_RNG},TRUE(),{E_RNG})'
                f'-AVERAGEIF({B_RNG},TRUE(),{C_RNG})^2),0)'
            )
            ws_calc.cell(r, 8).value  = f'=MIN({C_RNG})'
            ws_calc.cell(r, 9).value  = f'=MAX({C_RNG})'
            ws_calc.cell(r, 10).value = f'=IFERROR(AVERAGEIF({B_RNG},TRUE(),{C_RNG}),0)'
            ws_calc.cell(r, 11).value = f'=COUNTIF({B_RNG},TRUE())'

    print(f"  Calc 시트 수식 완료 ({n}행)")

    # ── Dashboard 수식 범위 업데이트 ────────────────────────────────
    old_n = 110
    for row in ws_dash.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = re.sub(
                    r'((?:Calc|Data)![A-Z]+2:[A-Z]+)' + str(old_n + 1),
                    lambda m: m.group(1) + str(n + 1),
                    cell.value
                )

    # ── ET_구조 시트 복사 ─────────────────────────────────────────────
    if 'ET_구조' in wb_et.sheetnames:
        ws_et_tree = wb_et['ET_구조']
        if 'ET_구조' in wb.sheetnames:
            del wb['ET_구조']
        ws_new = wb.create_sheet('ET_구조')
        for row in ws_et_tree.iter_rows():
            for cell in row:
                new_cell = ws_new.cell(row=cell.row, column=cell.column,
                                       value=cell.value)
                if cell.has_style:
                    new_cell.font      = copy(cell.font)
                    new_cell.fill      = copy(cell.fill)
                    new_cell.alignment = copy(cell.alignment)
        for col, dim in ws_et_tree.column_dimensions.items():
            ws_new.column_dimensions[col].width = dim.width
        print(f"  ET_구조 시트 복사 완료")

    # ── ET_Structure 시트 제목 업데이트 ──────────────────────────────
    if 'ET_Structure' in wb.sheetnames:
        ws_struct = wb['ET_Structure']
        if ws_struct['A1'].value and 'LOFW' in str(ws_struct['A1'].value):
            ws_struct['A1'] = ws_struct['A1'].value.replace(
                'LOFW (Loss of Feedwater)', accident_name
            )

    # ── 저장 + 차트 색상 패치 ─────────────────────────────────────────
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.save(output_name)
    _patch_chart_colors(output_name)

    print(f"  저장 완료: {output_name}")
    print(f"  사고유형: {accident_name} | 시나리오: {n}개 | CD기준: {CD_THRESHOLD}K")
    return str(output_name)


# ── 일괄 처리 ─────────────────────────────────────────────────────────
def run_all(et_paths: list, dash_path: Path, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    results, errors = {}, []

    for et_path in et_paths:
        et_path = Path(et_path)
        print(f"\n{'─'*70}")
        print(f"처리 중: {et_path.name}")
        print('─' * 70)
        try:
            out = process_one(et_path, dash_path, output_dir)
            results[et_path.name] = out
        except Exception as e:
            import traceback
            print(f"  오류: {e}")
            traceback.print_exc()
            errors.append((et_path.name, str(e)))

    print(f"\n{'='*70}")
    print(f"완료: {len(results)}개 성공 / {len(errors)}개 실패")
    for fname, out in results.items():
        print(f"   {fname} → {out}")
    if errors:
        print("\n실패 목록:")
        for fname, err in errors:
            print(f"   {fname}: {err}")
    print('=' * 70)


# ── CLI 진입점 ────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='ET result → PSA DET Dashboard 일괄 생성'
    )
    parser.add_argument(
        'inputs', nargs='*',
        help='ET result xlsx 파일 또는 폴더 (생략 시 현재 폴더의 모든 xlsx)'
    )
    parser.add_argument(
        '-t', '--template', required=True,
        help='Dashboard 템플릿 xlsx 경로 (필수)'
    )
    parser.add_argument(
        '-o', '--output', default='dashboard_results',
        help='결과 저장 폴더 (기본값: dashboard_results/)'
    )
    args = parser.parse_args()

    dash_path = Path(args.template)
    if not dash_path.exists():
        print(f"오류: 템플릿 파일을 찾을 수 없습니다 → {dash_path}")
        sys.exit(1)

    # ET result 파일 수집
    et_files = []
    if not args.inputs:
        et_files = [p for p in sorted(Path('.').glob('*.xlsx'))
                    if p != dash_path]
    else:
        for inp in args.inputs:
            p = Path(inp)
            if p.is_dir():
                et_files.extend(sorted(p.glob('*.xlsx')))
            elif p.suffix.lower() == '.xlsx' and p.exists():
                et_files.append(p)
            else:
                print(f"경고: '{inp}' 를 찾을 수 없거나 xlsx 파일이 아닙니다.")

    if not et_files:
        print("처리할 ET result 파일이 없습니다.")
        sys.exit(1)

    print("=" * 70)
    print("PSA DET Dashboard 생성기")
    print("=" * 70)
    print(f"템플릿: {dash_path}")
    print(f"대상 파일: {len(et_files)}개")

    run_all(et_files, dash_path, Path(args.output))


if __name__ == '__main__':
    main()
